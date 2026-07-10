"""Importa los estudiantes del Excel "LISTA ESTUDIANTES PROGRAMA SCRATCH Act.xlsx".

Qué hace:
  1. Lee la hoja de estudiantes (APELLIDOS, NOMBRES).
  2. Deduce el género de cada estudiante por el nombre.
  3. Asigna un código público (USUARIO):
        - Mujeres  -> tigre-rosado-N   (animal tigre, color rosado)
        - Hombres  -> leon-azul-N      (animal leon,  color azul)
     El número N crece por separado en cada género (1, 2, 3, ...).
  4. Escribe una columna nueva "USUARIO" en el Excel (guarda una copia _con_usuarios.xlsx).
  5. Inserta los estudiantes en la base de datos (tabla `estudiantes`) de forma
     idempotente, dentro del grupo "ESTUDIANTES NIVEL 0".

Uso:
    cd backend
    source venv/bin/activate
    python scripts/import_estudiantes_nivel0.py                # Excel + base de datos
    python scripts/import_estudiantes_nivel0.py --solo-excel   # solo actualiza el Excel
    python scripts/import_estudiantes_nivel0.py --dry-run      # muestra todo, no escribe nada

La conexión a la base de datos usa la misma configuración que la app
(variable DATABASE_URL / backend/.env). Ejecuta primero las migraciones
(`python scripts/create_tables.py`) si la tabla aún no existe.
"""

from __future__ import annotations

import argparse
import asyncio
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]          # .../backend
REPO_DIR = ROOT_DIR.parent                              # raíz del repo
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

EXCEL_IN = REPO_DIR / "LISTA ESTUDIANTES PROGRAMA SCRATCH Act.xlsx"
EXCEL_OUT = REPO_DIR / "LISTA ESTUDIANTES PROGRAMA SCRATCH Act_con_usuarios.xlsx"
GRUPO_NOMBRE = "ESTUDIANTES NIVEL 0"

# --- Configuración del código público ------------------------------------
ANIMAL_COLOR = {
    "femenino": ("tigre", "rosado"),
    "masculino": ("leon", "azul"),
}

# --- Diccionario de género por primer nombre ------------------------------
# Sin tildes y en minúsculas. Revísalo/edítalo si algún caso está mal:
# la clasificación automática de nombres nunca es perfecta.
NOMBRE_GENERO: dict[str, str] = {
    # Femenino
    "yesli": "F", "janeth": "F", "allison": "F", "keysha": "F", "aisha": "F",
    "shirley": "F", "eliana": "F", "scarlett": "F", "ruth": "F", "desire": "F",
    "arlett": "F", "brittany": "F", "domenica": "F", "dasha": "F", "jurybel": "F",
    "odalys": "F", "johana": "F", "kasandra": "F", "madison": "F", "lady": "F",
    "kerly": "F", "yelena": "F", "yahnia": "F", "elizabeth": "F", "michelle": "F",
    # Masculino
    "allan": "M", "wilmer": "M", "andy": "M", "keiner": "M", "leonarvis": "M",
    "snayder": "M", "yeyder": "M", "neymar": "M", "fausto": "M", "laionel": "M",
    "jesus": "M", "edgar": "M", "andres": "M", "alex": "M", "jose": "M",
    "juan": "M", "jeremy": "M", "samuel": "M", "luis": "M", "ezequiel": "M",
    "benjamin": "M", "leonel": "M", "xavier": "M", "keiler": "M", "thiago": "M",
    "justin": "M", "angel": "M", "natanael": "M", "isaac": "M", "kenny": "M",
}

GENERO_OPCION = {"F": "femenino", "M": "masculino"}


def sin_tildes(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def primer_nombre(nombres: str) -> str:
    return sin_tildes(nombres.strip().split()[0]).lower()


def deducir_genero(nombres: str) -> str:
    """Devuelve 'F' o 'M'. Lanza error si el nombre no está en el diccionario."""
    clave = primer_nombre(nombres)
    if clave not in NOMBRE_GENERO:
        raise KeyError(
            f"No sé el género de '{nombres}' (primer nombre '{clave}'). "
            f"Agrégalo a NOMBRE_GENERO en {__file__}."
        )
    return NOMBRE_GENERO[clave]


def leer_estudiantes() -> list[dict]:
    """Lee el Excel y devuelve una lista de dicts con apellidos, nombres, fila."""
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb.active
    estudiantes = []
    for fila in range(2, ws.max_row + 1):  # fila 1 = encabezados
        apellidos = ws.cell(row=fila, column=1).value
        nombres = ws.cell(row=fila, column=2).value
        if not apellidos and not nombres:
            continue
        estudiantes.append(
            {
                "fila": fila,
                "apellidos": (apellidos or "").strip(),
                "nombres": (nombres or "").strip(),
            }
        )
    return estudiantes


def asignar_codigos(estudiantes: list[dict]) -> None:
    """Agrega 'genero' y 'codigo' (USUARIO) a cada estudiante, en orden."""
    contador = {"F": 0, "M": 0}
    for est in estudiantes:
        g = deducir_genero(est["nombres"])
        contador[g] += 1
        animal, color = ANIMAL_COLOR[GENERO_OPCION[g]]
        est["genero"] = g
        est["codigo"] = f"{animal}-{color}-{contador[g]}"


def escribir_excel(estudiantes: list[dict], dry_run: bool) -> None:
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb.active
    col_usuario = ws.max_column + 1
    ws.cell(row=1, column=col_usuario, value="USUARIO")
    for est in estudiantes:
        ws.cell(row=est["fila"], column=col_usuario, value=est["codigo"])
    if dry_run:
        print(f"[dry-run] No se guardó el Excel (habría escrito {EXCEL_OUT.name}).")
        return
    wb.save(EXCEL_OUT)
    print(f"Excel actualizado -> {EXCEL_OUT}")


async def insertar_en_db(estudiantes: list[dict], dry_run: bool) -> None:
    from sqlalchemy import select

    from app.db.database import AsyncSessionLocal
    from app.models import Group, Student

    async with AsyncSessionLocal() as db:
        grupo = await db.scalar(select(Group).where(Group.nombre == GRUPO_NOMBRE))
        if grupo is None:
            grupo = Group(nombre=GRUPO_NOMBRE, descripcion="Estudiantes Nivel 0 programa Scratch")
            if not dry_run:
                db.add(grupo)
                await db.flush()
            print(f"Grupo creado: {GRUPO_NOMBRE}")
        else:
            print(f"Grupo existente: {GRUPO_NOMBRE}")

        creados, existentes = 0, 0
        for est in estudiantes:
            codigo = est["codigo"].strip().lower()
            ya = await db.scalar(select(Student).where(Student.codigo_publico == codigo))
            if ya is not None:
                existentes += 1
                continue
            creados += 1
            if not dry_run:
                # Nota: la tabla `estudiantes` no tiene columna de nombre en las
                # migraciones; el mapeo nombre<->codigo queda en el Excel (USUARIO).
                db.add(
                    Student(
                        codigo_publico=codigo,
                        grupo_id=grupo.id if grupo.id else None,
                        genero_opcion=GENERO_OPCION[est["genero"]],
                    )
                )

        if dry_run:
            print(f"[dry-run] Se insertarían {creados} estudiantes ({existentes} ya existían).")
            await db.rollback()
            return

        await db.commit()
        print(f"Base de datos: {creados} creados, {existentes} ya existían.")


def imprimir_tabla(estudiantes: list[dict]) -> None:
    print(f"\n{'USUARIO':<18} {'GÉNERO':<10} APELLIDOS Y NOMBRES")
    print("-" * 70)
    for est in estudiantes:
        genero = "Femenino" if est["genero"] == "F" else "Masculino"
        print(f"{est['codigo']:<18} {genero:<10} {est['apellidos']}, {est['nombres']}")
    n_f = sum(1 for e in estudiantes if e["genero"] == "F")
    n_m = sum(1 for e in estudiantes if e["genero"] == "M")
    print("-" * 70)
    print(f"Total: {len(estudiantes)}  |  Mujeres: {n_f}  |  Hombres: {n_m}\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa estudiantes Nivel 0 desde el Excel.")
    parser.add_argument("--solo-excel", action="store_true", help="Solo actualiza el Excel, no toca la base de datos.")
    parser.add_argument("--dry-run", action="store_true", help="Muestra todo sin escribir nada.")
    args = parser.parse_args()

    if not EXCEL_IN.exists():
        sys.exit(f"No se encontró el Excel: {EXCEL_IN}")

    estudiantes = leer_estudiantes()
    asignar_codigos(estudiantes)
    imprimir_tabla(estudiantes)

    escribir_excel(estudiantes, dry_run=args.dry_run)

    if not args.solo_excel:
        asyncio.run(insertar_en_db(estudiantes, dry_run=args.dry_run))


if __name__ == "__main__":
    main()
